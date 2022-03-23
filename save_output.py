import error_handling
import extract_member_forces as emf
import generate_mem_array_info as g_mem
import generate_joint_array_info as g_joi
import extract_joint_reactions as ejr
import extract_code_check as ecc
import generate_code_check_array_info as g_code
import csv
import openpyxl
from tkinter import *
from tkinter.ttk import *


# TODO: Add column headers to outputs

class RunProgram:

    def __init__(self, initial_window):
        self.initial_window = initial_window

    result_set_errors = []
    list_errors = []
    beam_errors = []
    load_errors = []
    joint_errors = []
    name_errors = []
    profile_errors = []
    ir_errors = []

    def run_member_forces(self, output_file_name, joint, member_set, out_format, input_file, beam_id, load_id):
        """
            Runs the full program. Saves all required outputs as selected by the user and saves either an xlsx file or a
            csv file.

                Parameters:
                    output_file_name (str):     name of file in which to save the data
                    joint (tuple):              list containing "ALL", "STA", or "END" for joint specification
                    member_set (list):          list of values corresponding to the index of requested member set
                    out_format (str):           .csv or .xlsx
                    input_file (str):           input file in which to sear for data
                    beam_id (list):             list containing tuples of beam spec (int), and user beam input (str)
                    load_id (list):             list containing tuples of load spec (int), and user load input (str)

                Returns:
                     .xlsx or .csv file of requested user inputs
            """
        self.result_set_errors.clear()
        self.list_errors.clear()
        self.beam_errors.clear()
        self.load_errors.clear()
        if out_format == '.xlsx':
            wb = openpyxl.Workbook()
            for items in range(len(member_set)):
                member_forces, end_index, first_useful_line \
                    = g_mem.ParseFileForData(items, input_file, member_set).get_member_force_list_info()
                member_forces, errors = emf.GenerateOutputArray(joint, items, member_forces,
                                                                beam_id, load_id).requested_member_force_array()
                sheet_name = 'Load Set ' + str(items + 1)
                sheet = wb.create_sheet(f'{sheet_name}')
                d_list = [list(k) + v for k, v in member_forces.items()]
                if len(member_forces) == 0:
                    self.result_set_errors.append(items + 1)

                if errors[0] or errors[1]:
                    self.list_errors.append(items + 1)
                    self.beam_errors.append(errors[0])
                    self.load_errors.append(errors[1])

                for row, key in enumerate(member_forces.items(), start=1):
                    for col, key in enumerate(d_list[row - 1], start=1):
                        sheet.cell(column=col, row=row, value=d_list[row - 1][col - 1])
            del wb['Sheet']
            wb.save(output_file_name)
        else:
            with open(output_file_name, 'w', newline=''):
                for items in range(len(member_set)):
                    member_forces, end_index, first_useful_line = \
                        g_mem.ParseFileForData(items, input_file, member_set).get_member_force_list_info()
                    member_forces, errors = emf.GenerateOutputArray(joint, items, member_forces,
                                                                    beam_id, load_id).requested_member_force_array()

                    with open(output_file_name, 'a', newline='') as a:
                        csv.writer(a).writerows((list(k) + v for k, v in member_forces.items()))

                    if len(member_forces) == 0:
                        self.result_set_errors.append(items + 1)

                    if errors[0] or errors[1]:
                        self.list_errors.append(items + 1)
                        self.beam_errors.append(errors[0])
                        self.load_errors.append(errors[1])

        if len(self.result_set_errors) or len(self.list_errors) > 0:
            error_set_list = list(set(self.list_errors + self.result_set_errors))
            error_handling.ErrorHandling(self.initial_window).item_not_found(error_set_list, self.beam_errors,
                                                                             self.load_errors, member_force=True)
        else:
            success_window = Toplevel(self.initial_window)
            success_window.geometry('200x120')
            success_window.title('Complete')
            success_window.resizable(False, False)
            success_window.grab_set()
            self.initial_window.eval(f'tk::PlaceWindow {str(success_window)} center')
            Label(success_window, text='Output Saved Successfully').pack()
            Button(success_window, text='OK', command=success_window.destroy).pack()

    def run_joint_reactions(self, output_file_name, joint_set, out_format, input_file, joint_id, load_id):
        """
            Runs the full program. Saves all required outputs as selected by the user and saves either an xlsx file or a
            csv file.

                Parameters:
                    output_file_name (str):     name of file in which to save the data
                    joint_set (list):          list of values corresponding to the index of requested joint set
                    out_format (str):           .csv or .xlsx
                    input_file (str):           input file in which to sear for data
                    joint_id (list):             list containing tuples of beam spec (int), and user beam input (str)
                    load_id (list):             list containing tuples of load spec (int), and user load input (str)

                Returns:
                     .xlsx or .csv file of requested user inputs
            """
        self.result_set_errors.clear()
        self.list_errors.clear()
        self.joint_errors.clear()
        self.load_errors.clear()
        if out_format == '.xlsx':
            wb = openpyxl.Workbook()
            for items in range(len(joint_set)):
                joint_reactions, end_index, first_useful_line \
                    = g_joi.ParseFileForData(items, input_file, joint_set).get_joint_reaction_list_info()
                joint_reactions, errors = ejr.GenerateOutputArray(items, joint_reactions,
                                                                  joint_id, load_id).requested_joint_reaction_dict()
                sheet_name = 'Load Set ' + str(items + 1)
                sheet = wb.create_sheet(f'{sheet_name}')
                d_list = [list(k) + v for k, v in joint_reactions.items()]
                if len(joint_reactions) == 0:
                    self.result_set_errors.append(items + 1)

                if errors[0] or errors[1]:
                    self.list_errors.append(items + 1)
                    self.joint_errors.append(errors[0])
                    self.load_errors.append(errors[1])

                for row, key in enumerate(joint_reactions.items(), start=1):
                    for col, key in enumerate(d_list[row - 1], start=1):
                        sheet.cell(column=col, row=row, value=d_list[row - 1][col - 1])
            del wb['Sheet']
            wb.save(output_file_name)
        else:
            with open(output_file_name, 'w', newline=''):
                for items in range(len(joint_set)):
                    joint_reactions, end_index, first_useful_line \
                        = g_joi.ParseFileForData(items, input_file, joint_set).get_joint_reaction_list_info()
                    joint_reactions, errors = ejr.GenerateOutputArray(items, joint_reactions,
                                                                      joint_id, load_id).requested_joint_reaction_dict()

                    with open(output_file_name, 'a', newline='') as a:
                        csv.writer(a).writerows((list(k) + v for k, v in joint_reactions.items()))

                    if len(joint_reactions) == 0:
                        self.result_set_errors.append(items + 1)

                    if errors[0] or errors[1]:
                        self.list_errors.append(items + 1)
                        self.joint_errors.append(errors[0])
                        self.load_errors.append(errors[1])

        if len(self.result_set_errors) or len(self.list_errors) > 0:
            error_set_list = list(set(self.list_errors + self.result_set_errors))
            error_handling.ErrorHandling(self.initial_window).item_not_found(error_set_list, self.joint_errors,
                                                                             self.load_errors, joint_reaction=True)
        else:
            success_window = Toplevel(self.initial_window)
            success_window.geometry('200x60')
            success_window.title('Complete')
            success_window.resizable(False, False)
            success_window.grab_set()
            self.initial_window.eval(f'tk::PlaceWindow {str(success_window)} center')
            Label(success_window, text='Output Saved Successfully').pack()
            Button(success_window, text='OK', command=success_window.destroy).pack()

    def run_code_check(self, output_file_name, code_set, out_format, input_file, name_id, profile_id, ir_range, fail_id,
                       sort_request, sort_order, reverse):
        # TODO: clean up save_output
        """
            Runs the full program. Saves all required outputs as selected by the user and saves either an xlsx file or a
            csv file.

                Parameters:
                    output_file_name (str):     name of file in which to save the data
                    joint_set (list):          list of values corresponding to the index of requested joint set
                    out_format (str):           .csv or .xlsx
                    input_file (str):           input file in which to sear for data
                    joint_id (list):             list containing tuples of beam spec (int), and user beam input (str)
                    load_id (list):             list containing tuples of load spec (int), and user load input (str)

                Returns:
                     .xlsx or .csv file of requested user inputs
            """
        self.result_set_errors.clear()
        self.list_errors.clear()
        self.name_errors.clear()
        self.profile_errors.clear()
        self.ir_errors.clear()
        if out_format == '.xlsx':
            wb = openpyxl.Workbook()
            for items in range(len(code_set)):
                code_check, end_index, first_useful_line \
                    = g_code.ParseFileForData(items, input_file, code_set).get_code_check_list_info()
                code_check, errors = ecc.GenerateOutputArray(code_check, items, name_id, profile_id, ir_range,
                                                                  fail_id).output_list(sort_request, sort_order, reverse)
                sheet_name = 'Load Set ' + str(items + 1)
                sheet = wb.create_sheet(f'{sheet_name}')
                if len(code_check) == 0:
                    self.result_set_errors.append(items + 1)

                if errors[0] or errors[1] or errors[2]:
                    self.list_errors.append(items + 1)
                    self.name_errors.append(errors[0])
                    self.profile_errors.append(errors[1])
                    self.ir_errors.append(errors[2])

                for row, key in enumerate(code_check, start=1):
                    for col, key in enumerate(code_check[row - 1], start=1):
                        sheet.cell(column=col, row=row, value=code_check[row - 1][col - 1])
            del wb['Sheet']
            wb.save(output_file_name)
        else:
            with open(output_file_name, 'w', newline=''):
                for items in range(len(code_set)):
                    code_check, end_index, first_useful_line \
                        = g_code.ParseFileForData(items, input_file, code_set).get_code_check_list_info()
                    code_check, errors = ecc.GenerateOutputArray(code_check, items, name_id, profile_id, ir_range,
                                                                      fail_id).output_list(sort_request, sort_order, reverse)

                    with open(output_file_name, 'a', newline='') as a:
                        csv.writer(a).writerows(code_check)

                    if len(code_check) == 0:
                        self.result_set_errors.append(items + 1)

                    if errors[0] or errors[1] or errors[2]:
                        self.list_errors.append(items + 1)
                        self.name_errors.append(errors[0])
                        self.profile_errors.append(errors[1])
                        self.ir_errors.append(errors[2])

        if len(self.result_set_errors) or len(self.list_errors) > 0:
            error_set_list = list(set(self.list_errors + self.result_set_errors))
            error_handling.ErrorHandling(self.initial_window).item_not_found(error_set_list, self.name_errors,
                                                                             self.profile_errors, ir_errors=self.ir_errors, code_check=True)
        else:
            success_window = Toplevel(self.initial_window)
            success_window.geometry('200x60')
            success_window.title('Complete')
            success_window.resizable(False, False)
            success_window.grab_set()
            self.initial_window.eval(f'tk::PlaceWindow {str(success_window)} center')
            Label(success_window, text='Output Saved Successfully').pack()
            Button(success_window, text='OK', command=success_window.destroy).pack()
