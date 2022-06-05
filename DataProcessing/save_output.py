"""
This module saves the requested output as a .csv or .xlsx file. Iteratively processes all result sets in the
selected results tree one at a time.
"""
import csv
from tkinter import *
from tkinter.ttk import *
import openpyxl
from Tools import config
from DataProcessing import extract_member_forces as emf
from DataProcessing.parse_file_for_input_data import ParseFileForData
from DataProcessing import extract_joint_reactions as ejr
from DataProcessing import extract_code_check as ecc
from error_handling import ErrorHandling


class RunProgram:
    """
    Runs the program and generates and output file with the requested results

    :param tab_name: active tab name
    :param output_format: user requested output format (.csv, .xlsx)
    :param output_file_name: name of output file
    :param result_set_index: result set being processed
    :param initial_window: active window
    """
    def __init__(self, tab_name, output_format, output_file_name, result_set_index, initial_window=None):
        self.initial_window = initial_window
        self.output_format = output_format
        self.tab_name = tab_name
        self.output_file_name = output_file_name
        self.result_set_index = result_set_index
        if self.output_format == '.xlsx':
            self.generate_xlsx()
        else:
            self.generate_csv()

    def build_output(self, items):
        """
        builds dictionary or list of results which meet user input requirements.

        :param items: current item being processed
        :return: dictionary or list of parsed results corresponding the the specific item being saved.
        """
        extracted_result_list, _, _ = ParseFileForData(items, self.tab_name).get_result_list_info()
        if self.tab_name == 'Member Force':
            parsed_results = emf.GenerateOutputArray(self.tab_name, items, extracted_result_list,
                                                     ).requested_member_force_array()
        elif self.tab_name == 'Joint Reaction':
            parsed_results = ejr.GenerateOutputArray(self.tab_name, items,
                                                     extracted_result_list).requested_joint_reaction_dict()
        else:
            parsed_results = ecc.GenerateOutputArray(self.tab_name, items,
                                                     extracted_result_list).output_list()
        return parsed_results

    def generate_xlsx(self):
        """
        saves the requested outputs as a .xlsx file
        """
        try:
            wb = openpyxl.Workbook()
            for items in range(len(self.result_set_index)):
                parsed_results = self.build_output(items)
                sheet_name = 'Result Set ' + str(items + 1)
                sheet = wb.create_sheet(f'{sheet_name}')
                for col, val in enumerate(config.result_configuration_parameters[self.tab_name]['Headings'], start=1):
                    sheet.cell(column=col, row=1, value=val)
                if self.tab_name == 'Code Check':
                    for row, key in enumerate(parsed_results, start=2):
                        for col, key in enumerate(parsed_results[row - 2], start=1):
                            sheet.cell(column=col, row=row, value=parsed_results[row - 2][col - 1])
                else:
                    d_list = [list(k) + v for k, v in parsed_results.items()]
                    for row, key in enumerate(parsed_results.items(), start=2):
                        for col, key in enumerate(d_list[row - 2], start=1):
                            sheet.cell(column=col, row=row, value=d_list[row - 2][col - 1])
            del wb['Sheet']
            wb.save(self.output_file_name)
            self.display_success_or_error()
        except PermissionError as e:
            print(e)
            ErrorHandling(self.initial_window).file_already_open(e)

    def generate_csv(self):
        """
        saves requested results as .csv file
        """
        try:
            with open(self.output_file_name, 'w', newline='') as w:
                csv.writer(w).writerow(config.result_configuration_parameters[self.tab_name]['Headings'])
                w.close()
                for items in range(len(self.result_set_index)):
                    parsed_results = self.build_output(items)
                    print(parsed_results)
                    with open(self.output_file_name, 'a', newline='') as a:
                        if self.tab_name == 'Code Check':
                            csv.writer(a).writerows(parsed_results)
                        else:
                            csv.writer(a).writerows((list(k) + v for k, v in parsed_results.items()))
            self.display_success_or_error()
        except PermissionError as e:
            print(e)
            ErrorHandling(self.initial_window).file_already_open(e)

    def display_success_or_error(self):
        """
        displays popup window when output is successfully saved.
        """
        success_window = Toplevel(self.initial_window)
        success_window.geometry('200x60')
        success_window.title('Complete')
        success_window.resizable(False, False)
        success_window.grab_set()
        self.initial_window.eval(f'tk::PlaceWindow {str(success_window)} center')
        Label(success_window, text='Output Saved Successfully').pack()
        Button(success_window, text='OK', command=success_window.destroy).pack()
